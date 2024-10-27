import openpyxl

def clear_excel_data():
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook("network_stats.xlsx")
        sheet = workbook.active

        # Clear all rows except the header (first row)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None  # Clear the cell's value

        # Save the workbook after clearing data
        workbook.save("network_stats.xlsx")
        print("Excel data cleared, except for headers.")
    except FileNotFoundError:
        print("No Excel file found to clear.")

if __name__ == "__main__":
    # Run the function to clear the data
    clear_excel_data()
